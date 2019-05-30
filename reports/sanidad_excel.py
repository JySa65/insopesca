import time
import calendar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, colors, Alignment, Border, Side, PatternFill
from django.views.generic import View
from django.http import HttpResponse, FileResponse
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from isoweek import Week
from sanidad.models import TypeCompany, Company
from utils.date_insopesca import DateInsopesca
from utils.get_data_report import get_type_company_report
FILENAME = f'{settings.MEDIA_ROOT}/reports_excel.xlsx'


class InpectionsCompany(View):
    def get(self, request, *args, **kwargs):
        year = request.GET.get('year', '')
        if year == '':
            year = datetime.now().year
        year = datetime(int(year), 1, 1).year
        companys, type_companys, total = self.get_data(year)
        self.form(companys, type_companys, total, year)
        fs = FileSystemStorage()
        # return HttpResponse("si")
        with fs.open(FILENAME) as xlsx:
            response = HttpResponse(xlsx)
            response['Content-type'] = 'application/ms-excel'
            response['Content-Disposition'] = f'inline; filename="reporte_excel_{year}.xlsx"'
        return response

    def form(self, companys, type_companys, total_rows, year):
        wb = Workbook()
        ws = wb.active

        # title
        thin = Side(border_style="thin", color="010204")

        ws.row_dimensions[1].height = 40
        ws.cell(row=1, column=1).fill = PatternFill(
            start_color='25d5f2', end_color='25d5f2', fill_type="solid")
        ws.merge_cells(f'A1:{get_column_letter(len(type_companys) + 4)}1')
        ws.cell(row=1, column=1).value = "Inspecciones Sanitarias Insopesca"
        ws.cell(row=1, column=1).border = Border(bottom=thin)
        ws.cell(row=1, column=len(type_companys) +
                3).border = Border(right=thin)
        ws.cell(row=1, column=1).font = Font(
            size=30, name='Arial', color='000000')
        ws.cell(row=1, column=1).alignment = Alignment(
            horizontal="center", vertical="center")

        #  week
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 20
        ws.column_dimensions['A'].width = 20
        ws.merge_cells('A3:A4')

        ws.cell(row=3, column=1).fill = PatternFill(
            start_color='bbbcbd', end_color='bbbcbd', fill_type="solid")
        ws.cell(row=3, column=1).value = "MES"
        ws.cell(row=3, column=1).border = Border(
            top=thin, left=thin, right=thin)
        ws.cell(row=4, column=1).border = Border(bottom=thin)
        ws.cell(row=3, column=1).font = Font(
            size=14, name='Arial', bold=True)
        ws.cell(row=3, column=1).alignment = Alignment(
            horizontal="center", vertical="center")

        ws.column_dimensions['B'].width = 15
        ws.merge_cells('B3:B4')
        ws.cell(row=3, column=2).fill = PatternFill(
            start_color='bbbcbd', end_color='bbbcbd', fill_type="solid")
        ws.cell(row=3, column=2).value = "Semana"
        ws.cell(row=3, column=2).border = Border(
            top=thin, left=thin, right=thin)
        ws.cell(row=4, column=2).border = Border(bottom=thin)
        ws.cell(row=3, column=2).font = Font(
            size=14, name='Arial', bold=True)
        ws.cell(row=3, column=2).alignment = Alignment(
            horizontal="center", vertical="center")

        ws.column_dimensions['C'].width = 30
        ws.merge_cells('C3:C4')
        ws.cell(row=3, column=3).fill = PatternFill(
            start_color='bbbcbd', end_color='bbbcbd', fill_type="solid")
        ws.cell(row=3, column=3).value = "Rango de Fecha"
        ws.cell(row=3, column=3).border = Border(
            top=thin, left=thin, right=thin)
        ws.cell(row=4, column=3).border = Border(bottom=thin)
        ws.cell(row=3, column=3).font = Font(
            size=14, name='Arial', bold=True)
        ws.cell(row=3, column=3).alignment = Alignment(
            horizontal="center", vertical="center")

        ws.merge_cells(f'D3:{get_column_letter(len(type_companys) + 4)}3')
        ws.cell(row=3, column=4).fill = PatternFill(
            start_color='bbbcbd', end_color='bbbcbd', fill_type="solid")
        ws.cell(row=3, column=4).value = "Tipo De Compañia"
        ws.cell(row=3, column=4).border = Border(
            bottom=thin, top=thin, left=thin)
        ws.cell(row=3, column=len(type_companys) +
                4).border = Border(right=thin)
        ws.cell(row=3, column=4).font = Font(
            size=14, name='Arial', bold=True)
        ws.cell(row=3, column=4).alignment = Alignment(
            horizontal="center", vertical="center")

        col_total = len(type_companys) + 4
        ws.cell(row=4, column=col_total).fill = PatternFill(
            start_color='bbbcbd', end_color='bbbcbd', fill_type="solid")
        ws.cell(row=4, column=col_total).value = "Total"
        ws.cell(row=4, column=col_total).border = Border(
            bottom=thin, top=thin, left=thin, right=thin)
        ws.cell(row=4, column=col_total).font = Font(
            size=14, name='Arial', bold=True)
        ws.cell(row=4, column=col_total).alignment = Alignment(
            horizontal="center", vertical="center")

        for key, type_company in enumerate(type_companys):
            col = key + 4
            row = 4
            value = type_company['name'].title()
            ws.column_dimensions[get_column_letter(col)].width = len(value) * 2
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color='bbbcbd', end_color='bbbcbd', fill_type="solid")
            ws.cell(row=row, column=col).value = value
            ws.cell(row=row, column=col).border = Border(
                bottom=thin, top=thin, left=thin, right=thin)
            ws.cell(row=row, column=col).font = Font(
                size=12, name='Arial', bold=True)
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal="center", vertical="center")

        for k, company in enumerate(companys):
            row = k + 5
            month = company['month']
            ws.cell(row=row, column=1).value = self.get_month()[month-1]
            ws.cell(row=row, column=1).border = Border(
                bottom=thin, top=thin, left=thin, right=thin)
            ws.cell(row=row, column=1).font = Font(
                size=11, name='Arial', bold=True)
            ws.cell(row=row, column=1).alignment = Alignment(
                horizontal="center", vertical="center")

            # week number
            ws.cell(row=row, column=2).value = k + 1
            ws.cell(row=row, column=2).border = Border(
                bottom=thin, top=thin, left=thin, right=thin)
            ws.cell(row=row, column=2).font = Font(
                size=11, name='Arial')
            ws.cell(row=row, column=2).alignment = Alignment(
                horizontal="center", vertical="center")

            # range date
            ws.cell(row=row, column=3).value = company['range_week']
            ws.cell(row=row, column=3).border = Border(
                bottom=thin, top=thin, left=thin, right=thin)
            ws.cell(row=row, column=3).font = Font(
                size=11, name='Arial')
            ws.cell(row=row, column=3).alignment = Alignment(
                horizontal="center", vertical="center")

            # total
            col_total = len(type_companys) + 4
            total_col = company['total_col'] if company['total_col'] != 0 else ''
            ws.cell(row=row, column=col_total).value = total_col
            ws.cell(row=row, column=col_total).border = Border(
                bottom=thin, top=thin, left=thin, right=thin)
            ws.cell(row=row, column=col_total).font = Font(
                size=11, name='Arial', bold=True)
            ws.cell(row=row, column=col_total).alignment = Alignment(
                horizontal="center", vertical="center")

            for key, total in enumerate(company['inspections_total']):
                col = (key) + 4
                value = total
                if value == 0:
                    value = ""
                ws.cell(row=row, column=col).value = value
                ws.cell(row=row, column=col).border = Border(
                    bottom=thin, top=thin, left=thin, right=thin)
                ws.cell(row=row, column=col).font = Font(
                    size=11, name='Arial')
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal="center", vertical="center")

        rowe = row + 1
        ws.row_dimensions[rowe].height = 30
        ws.cell(row=rowe, column=1).fill = PatternFill(
            start_color='bbbcbd', end_color='bbbcbd', fill_type="solid")
        ws.merge_cells(
            f'A{rowe}:{get_column_letter(len(type_companys) + 3)}{rowe}')
        ws.cell(
            row=rowe, column=1).value = f"Total Inspecciones Del Año {year}"
        ws.cell(row=rowe, column=1).border = Border(bottom=thin)
        ws.cell(row=rowe, column=len(type_companys) +
                3).border = Border(right=thin)
        ws.cell(row=rowe, column=1).font = Font(
            size=20, name='Arial', color='000000')
        ws.cell(row=rowe, column=1).alignment = Alignment(
            horizontal="center", vertical="center")

        cole = len(type_companys) + 4
        ws.cell(row=rowe, column=cole).value = total_rows
        ws.cell(row=rowe, column=cole).border = Border(bottom=thin, right=thin)
        ws.cell(row=rowe, column=cole).font = Font(
            size=14, name='Arial', bold=True)
        ws.cell(row=rowe, column=cole).alignment = Alignment(
            horizontal="center", vertical="center")

        wb.save(FILENAME)
        return True

    def get_data(self, year):
        type_comanys = [dict(pk=i.pk, name=i.name)
                        for i in TypeCompany.objects.all()]
        data = []
        _range = Week.last_week_of_year(year)[1]
        for i in range(_range):
            number = i+1
            date_list = Week(year, number).days()
            month = self.get_month_from_week(date_list)
            first, last = DateInsopesca(
                year, number).get_date_range_from_week()
            data.append(dict(
                range_week=f'{first} a {last}',
                inspections_total=[],
                month=month,
                total_col=0))
            for type_comany in type_comanys:
                inspection_total = 0
                companys = Company.objects.filter(
                    type_company=type_comany['pk'])
                for company in companys:
                    inspections = company.get_inspections(
                        f'{year}-{number}', f'{year}-{number}')
                    inspection_total += len(inspections)
                data[i]['inspections_total'].append(inspection_total)
                data[i]['total_col'] += inspection_total

        total = 0
        for tot in data:
            total += tot['total_col']
        return data, type_comanys, total

    def get_month_from_week(self, date_list):
        month = 0
        for date in date_list:
            if date.month != month:
                month = date.month
        return month

    def get_month(self):
        return [
            "ENERO",
            "FEBRERO",
            "MARZO",
            "ABRIL",
            "MAYO",
            "JUNIO",
            "JULIO",
            "AGOSTO",
            "SEPTIEMBRE",
            "OCTUBRE",
            "NOVIEMBRE",
            "DICIEMBRE"
        ]
